"""
Tests for scraper/storage module.
Tests CSV, SQLite, JSON, and Excel storage backends.
"""

import json
import os
import pytest
import sqlite3
import tempfile
from pathlib import Path

# Import storage modules
from scraper.storage import csv_file, json_file, sqlite, excel


class TestCSVStorage:
    """Tests for CSV storage backend."""

    def test_save_creates_csv_file(self):
        """save() should create a CSV file with the data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"name": "test", "value": 123, "url": "https://example.com"}
            result = csv_file.save(data, "test_csv", output_dir=tmpdir)
            
            assert Path(result).exists()
            assert result.endswith(".csv")

    def test_save_appends_to_existing_csv(self):
        """save() should append rows to existing CSV."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data1 = {"name": "first", "value": 1}
            data2 = {"name": "second", "value": 2}
            
            csv_file.save(data1, "test_append", output_dir=tmpdir)
            csv_file.save(data2, "test_append", output_dir=tmpdir)
            
            csv_path = Path(tmpdir) / "test_append.csv"
            content = csv_path.read_text()
            
            # Should have header + 2 data rows
            lines = content.strip().split("\n")
            assert len(lines) == 3  # header + 2 rows
            assert "first" in lines[1]
            assert "second" in lines[2]

    def test_save_creates_header_on_first_write(self):
        """First save should create CSV with header."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"title": "Hello", "count": 42}
            csv_file.save(data, "header_test", output_dir=tmpdir)

            csv_path = Path(tmpdir) / "header_test.csv"
            content = csv_path.read_text()

            assert "title" in content
            assert "count" in content

    def test_save_handles_special_characters(self):
        """save() should handle special characters in data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"text": "Hello\nWorld", "quote": 'He said "hi"'}
            result = csv_file.save(data, "special_chars", output_dir=tmpdir)
            
            assert Path(result).exists()


class TestJSONStorage:
    """Tests for JSON file storage backend."""

    def test_save_creates_json_file(self):
        """save() should create a JSON file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"key": "value", "number": 42}
            result = json_file.save(data, "test_json", output_dir=tmpdir)
            
            assert Path(result).exists()
            assert result.endswith(".json")

    def test_save_writes_correct_data(self):
        """save() should write correct data to JSON file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"name": "test", "items": [1, 2, 3]}
            result = json_file.save(data, "data_test", output_dir=tmpdir)
            
            with open(result, "r") as f:
                loaded = json.load(f)
            
            assert loaded["name"] == "test"
            assert loaded["items"] == [1, 2, 3]

    def test_save_overwrites_existing_file(self):
        """save() should overwrite existing JSON file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            json_file.save({"old": "data"}, "overwrite_test", output_dir=tmpdir)
            json_file.save({"new": "data"}, "overwrite_test", output_dir=tmpdir)
            
            json_path = Path(tmpdir) / "overwrite_test.json"
            loaded = json.loads(json_path.read_text())
            
            assert "new" in loaded
            assert "old" not in loaded


class TestSQLiteStorage:
    """Tests for SQLite storage backend."""

    def test_save_creates_database(self):
        """save() should create a SQLite database."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"url": "https://example.com", "title": "Test Page"}
            result = sqlite.save(data, "test_directive", output_dir=tmpdir)
            
            assert Path(result).exists()
            assert result.endswith(".db")

    def test_save_inserts_record(self):
        """save() should insert a record into the database."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"url": "https://test.com", "content": "Hello"}
            result = sqlite.save(data, "my_directive", output_dir=tmpdir)
            
            # Verify by reading the database directly
            conn = sqlite3.connect(result)
            cursor = conn.execute("SELECT * FROM scrapes")
            rows = cursor.fetchall()
            conn.close()
            
            assert len(rows) >= 1
            assert rows[0][1] == "my_directive"  # directive column

    def test_save_handles_special_characters(self):
        """save() should handle special characters in data."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"url": "https://test.com", "content": "Hello\nWorld"}
            result = sqlite.save(data, "special", output_dir=tmpdir)
            
            assert Path(result).exists()

    def test_save_creates_table_if_not_exists(self):
        """save() should create the scrapes table if it doesn't exist."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"url": "https://test.com"}
            result = sqlite.save(data, "new_directive", output_dir=tmpdir)
            
            # Verify table exists
            conn = sqlite3.connect(result)
            cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='scrapes'")
            table_exists = cursor.fetchone() is not None
            conn.close()
            
            assert table_exists

    def test_save_creates_indexes(self):
        """save() should create indexes for performance."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"url": "https://test.com"}
            sqlite.save(data, "index_test", output_dir=tmpdir)
            
            # Verify indexes exist
            conn = sqlite3.connect(Path(tmpdir) / "scrapit.db")
            cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='index'")
            indexes = [row[0] for row in cursor.fetchall()]
            conn.close()
            
            assert "idx_directive" in indexes
            assert "idx_url" in indexes


class TestExcelStorage:
    """Tests for Excel storage backend."""

    def test_save_creates_xlsx_file(self):
        """save() should create an Excel file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"name": "test", "value": 100}
            result = excel.save(data, "test_excel", output_dir=tmpdir)
            
            assert Path(result).exists()
            assert result.endswith(".xlsx")

    def test_save_creates_headers_on_first_write(self):
        """First save should create headers in the Excel file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            data = {"col_a": "A", "col_b": "B"}
            excel.save(data, "header_test", output_dir=tmpdir)
            
            # Load and verify
            from openpyxl import load_workbook
            wb = load_workbook(Path(tmpdir) / "header_test.xlsx")
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            assert "col_a" in headers
            assert "col_b" in headers

    def test_save_appends_rows_to_existing(self):
        """save() should append rows to existing Excel file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel.save({"id": 1, "name": "First"}, "append_test", output_dir=tmpdir)
            excel.save({"id": 2, "name": "Second"}, "append_test", output_dir=tmpdir)
            
            from openpyxl import load_workbook
            wb = load_workbook(Path(tmpdir) / "append_test.xlsx")
            ws = wb.active
            
            # Should have 3 rows: 1 header + 2 data
            assert ws.max_row == 3  # header + 2 rows

    def test_save_adds_new_columns(self):
        """save() should add new columns if data has new keys."""
        with tempfile.TemporaryDirectory() as tmpdir:
            excel.save({"a": 1, "b": 2}, "new_cols", output_dir=tmpdir)
            excel.save({"a": 3, "b": 4, "c": 5}, "new_cols", output_dir=tmpdir)
            
            from openpyxl import load_workbook
            wb = load_workbook(Path(tmpdir) / "new_cols.xlsx")
            ws = wb.active
            
            headers = [cell.value for cell in ws[1]]
            assert "c" in headers


class TestStorageIntegration:
    """Integration tests for storage backends."""

    def test_csv_json_excel_handle_same_data_format(self):
        """CSV, JSON, and Excel should handle the same data format."""
        test_data = {
            "url": "https://example.com",
            "title": "Test Page",
            "timestamp": "2024-01-01T00:00:00",
            "items": ["a", "b", "c"]
        }
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # All should not raise exceptions
            csv_file.save(test_data.copy(), "integration", output_dir=tmpdir)
            json_file.save(test_data.copy(), "integration", output_dir=tmpdir)
            excel.save(test_data.copy(), "integration", output_dir=tmpdir)
            
            # Verify files were created
            assert (Path(tmpdir) / "integration.csv").exists()
            assert (Path(tmpdir) / "integration.json").exists()
            assert (Path(tmpdir) / "integration.xlsx").exists()
            
            # SQLite should also work (with its own db file)
            sqlite.save(test_data.copy(), "integration", output_dir=tmpdir)
            assert (Path(tmpdir) / "scrapit.db").exists()
